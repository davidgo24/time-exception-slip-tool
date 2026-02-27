import os
import io
import csv
import json
import base64
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Tuple

from flask import (
    Flask, render_template, request, jsonify, send_file, session
)
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch

logging.getLogger("PyPDF2").setLevel(logging.CRITICAL)
logging.getLogger("PyPDF2").propagate = False

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "montebello-transit-ot-2026")

TEMPLATE_PDF = os.path.join(os.path.dirname(__file__), "OT_Time_Exception_Slip_Sample.pdf")
DEPT_CODE = "910"
DATE_FMT_OUTPUT = "%m-%d-%y"
PAGE_W, PAGE_H = letter  # 612 x 792

with open(TEMPLATE_PDF, "rb") as _f:
    _TEMPLATE_BYTES = _f.read()

# ---------------------------------------------------------------------------
# Field coordinate map (extracted from PDF annotations)
# Each entry: (x, y, w, h, font_size, align)
#   align: "left" or "center"
# ---------------------------------------------------------------------------
FIELD_COORDS = {
    "Employee Name": (99.00, 711.72, 201.60, 16.92, 11, "left"),
    "Dept":          (345.84, 711.72, 45.60, 16.92, 11, "center"),
    "Ending Date":   (459.24, 711.72, 104.76, 16.92, 10, "left"),
    "Employee":      (76.32, 681.36, 65.52, 16.92, 10, "left"),

    # OT date columns (rows 1-5)
    "Dates 1_2": (18.48, 365.16, 80.76, 16.92, 7, "left"),
    "Dates 2_2": (18.48, 347.04, 80.76, 16.92, 7, "left"),
    "Dates 3_2": (18.48, 328.92, 80.76, 16.92, 7, "left"),
    "Dates 4_2": (18.48, 310.80, 80.76, 16.92, 7, "left"),
    "Dates 5_2": (18.48, 292.68, 80.76, 16.92, 7, "left"),

    # OT 1.0 (rows 1-5 + total row 6)
    "OT1": (143.64, 366.78, 18.33, 15.30, 7, "center"),
    "OT2": (143.64, 348.66, 18.33, 15.30, 7, "center"),
    "OT3": (143.64, 330.54, 18.33, 15.30, 7, "center"),
    "OT4": (143.64, 312.42, 18.33, 15.30, 7, "center"),
    "OT5": (143.64, 294.30, 18.33, 15.30, 7, "center"),
    "OT6": (143.64, 275.58, 18.33, 15.30, 7, "center"),

    # OT 1.5 (rows 1-5 + total row 6)
    "OTH1": (166.31, 366.78, 18.33, 15.30, 7, "center"),
    "OTH2": (166.31, 348.66, 18.33, 15.30, 7, "center"),
    "OTH3": (166.31, 330.54, 18.33, 15.30, 7, "center"),
    "OTH4": (166.31, 312.42, 18.33, 15.30, 7, "center"),
    "OTH5": (166.31, 294.30, 18.33, 15.30, 7, "center"),
    "OTH6": (166.31, 275.58, 18.33, 15.30, 7, "center"),

    # CTE 1.0 (rows 1-5 + total row 6)
    "CTE1": (189.05, 366.78, 18.33, 15.30, 7, "center"),
    "CTE2": (189.05, 348.66, 18.33, 15.30, 7, "center"),
    "CTE3": (189.05, 330.54, 18.33, 15.30, 7, "center"),
    "CTE4": (189.05, 312.42, 18.33, 15.30, 7, "center"),
    "CTE5": (189.05, 294.30, 18.33, 15.30, 7, "center"),
    "CTE6": (189.05, 275.58, 18.33, 15.30, 7, "center"),

    # CTE 1.5 (rows 1-5 + total row 6)
    "CTEH1": (211.98, 366.78, 18.33, 15.30, 7, "center"),
    "CTEH2": (211.98, 348.66, 18.33, 15.30, 7, "center"),
    "CTEH3": (211.98, 330.54, 18.33, 15.30, 7, "center"),
    "CTEH4": (211.98, 312.42, 18.33, 15.30, 7, "center"),
    "CTEH5": (211.98, 294.30, 18.33, 15.30, 7, "center"),
    "CTEH6": (211.98, 275.58, 18.33, 15.30, 7, "center"),

    # Grand total hours
    "HTOT1": (302.43, 275.58, 18.33, 15.30, 7, "center"),
}

OT_ROW_FIELDS = {
    "dates": ["Dates 1_2", "Dates 2_2", "Dates 3_2", "Dates 4_2", "Dates 5_2"],
    "ot10":  ["OT1", "OT2", "OT3", "OT4", "OT5"],
    "ot15":  ["OTH1", "OTH2", "OTH3", "OTH4", "OTH5"],
    "cte10": ["CTE1", "CTE2", "CTE3", "CTE4", "CTE5"],
    "cte15": ["CTEH1", "CTEH2", "CTEH3", "CTEH4", "CTEH5"],
}

OT_TOTAL_FIELDS = {
    "ot10": "OT6",
    "ot15": "OTH6",
    "cte10": "CTE6",
    "cte15": "CTEH6",
}

HOURS_TOTAL_FIELD = "HTOT1"

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def parse_date_flexible(s: str) -> datetime:
    s = s.strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"Cannot parse date: {s}")


def pay_period_weeks(ending_date_str: str) -> Tuple[Tuple[datetime, datetime], Tuple[datetime, datetime]]:
    end = parse_date_flexible(ending_date_str)
    # ending date is always a Saturday; pay period is 14 days Sun-Sat, Sun-Sat
    wk1_start = end - timedelta(days=13)  # Sunday
    wk1_end = end - timedelta(days=7)     # Saturday
    wk2_start = end - timedelta(days=6)   # Sunday
    wk2_end = end                         # Saturday
    return (wk1_start, wk1_end), (wk2_start, wk2_end)


def format_date_short(dt: datetime) -> str:
    return f"{dt.month}/{dt.day}"


def read_csv_bytes(file_bytes: bytes) -> List[Dict[str, str]]:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    for enc in encodings:
        try:
            text = file_bytes.decode(enc)
            reader = csv.DictReader(io.StringIO(text))
            return list(reader)
        except (UnicodeDecodeError, csv.Error):
            continue
    text = file_bytes.decode("utf-8", errors="replace")
    return list(csv.DictReader(io.StringIO(text)))


def get_field(row: dict, keys: list) -> str:
    for k in keys:
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def parse_employees_from_csv(file_bytes: bytes) -> List[Dict[str, str]]:
    rows = read_csv_bytes(file_bytes)
    employees = []
    for row in rows:
        last = get_field(row, ["LastName", "Last", "Last_Name", "Surname"])
        first = get_field(row, ["FirstName", "First", "First_Name", "GivenName"])
        emp_no = get_field(row, ["EmployeeNumber", "Employee #", "EmpNo", "EmployeeID", "Employee_Id"])
        if last or first:
            employees.append({
                "last": last,
                "first": first,
                "emp_no": emp_no,
            })
    employees.sort(key=lambda e: (e["last"].lower(), e["first"].lower()))
    return employees


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

def _create_overlay(values: dict) -> bytes:
    """Create a transparent PDF overlay with text drawn at field coordinates."""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)

    for field_name, text in values.items():
        if not text or field_name not in FIELD_COORDS:
            continue
        x, y, w, h, font_size, align = FIELD_COORDS[field_name]

        c.setFont("Helvetica", font_size)
        text_y = y + (h - font_size) / 2 + 1  # vertically center

        if align == "center":
            text_x = x + w / 2
            c.drawCentredString(text_x, text_y, str(text))
        else:
            text_x = x + 2  # small left padding
            c.drawString(text_x, text_y, str(text))

    c.save()
    return buf.getvalue()


def fill_single_pdf(employee: dict, pp_end: str, ot_data: dict = None) -> bytes:
    """Fill a single Time Exception Slip PDF using a reportlab text overlay."""
    combined_name = f"{employee['last']}, {employee['first']}".strip(", ").strip()
    end_dt = parse_date_flexible(pp_end)
    pp_end_formatted = end_dt.strftime(DATE_FMT_OUTPUT)

    values = {
        "Employee Name": combined_name,
        "Dept": DEPT_CODE,
        "Ending Date": pp_end_formatted,
        "Employee": employee["emp_no"],
    }

    if ot_data:
        (wk1_start, wk1_end), (wk2_start, wk2_end) = pay_period_weeks(pp_end)
        weeks = _aggregate_ot_by_week(ot_data, wk1_start, wk1_end, wk2_start, wk2_end)

        grand_total = 0.0
        for week_idx, week in enumerate(weeks):
            if not week["has_data"]:
                continue
            row_idx = week_idx

            values[OT_ROW_FIELDS["dates"][row_idx]] = week["dates_str"]

            for cat_key in ("ot10", "ot15", "cte10", "cte15"):
                hrs = week[cat_key]
                if hrs > 0:
                    values[OT_ROW_FIELDS[cat_key][row_idx]] = _fmt_hours(hrs)

            grand_total += week["row_total"]

        for cat_key in ("ot10", "ot15", "cte10", "cte15"):
            total = sum(w[cat_key] for w in weeks)
            if total > 0:
                values[OT_TOTAL_FIELDS[cat_key]] = _fmt_hours(total)

        if grand_total > 0:
            values[HOURS_TOTAL_FIELD] = _fmt_hours(grand_total)

    overlay_bytes = _create_overlay(values)

    template_reader = PdfReader(io.BytesIO(_TEMPLATE_BYTES))
    overlay_reader = PdfReader(io.BytesIO(overlay_bytes))

    template_page = template_reader.pages[0]
    overlay_page = overlay_reader.pages[0]
    template_page.merge_page(overlay_page)

    # Strip form field annotations to prevent name collisions when merging
    # multiple copies into a single binder PDF. Text is drawn via overlay.
    if "/Annots" in template_page:
        del template_page["/Annots"]

    writer = PdfWriter()
    writer.add_page(template_page)

    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _aggregate_ot_by_week(ot_data: dict, wk1_start, wk1_end, wk2_start, wk2_end) -> list:
    """Group OT entries into week 1 and week 2, summing hours by category."""
    weeks = [
        {"ot10": 0.0, "ot15": 0.0, "cte10": 0.0, "cte15": 0.0, "dates": set(), "has_data": False, "row_total": 0.0},
        {"ot10": 0.0, "ot15": 0.0, "cte10": 0.0, "cte15": 0.0, "dates": set(), "has_data": False, "row_total": 0.0},
    ]

    entries = ot_data.get("entries", [])
    for entry in entries:
        try:
            dt = parse_date_flexible(entry["date"])
        except ValueError:
            continue

        cat = entry.get("category", "")
        hours = float(entry.get("hours", 0))
        if hours <= 0:
            continue

        if wk1_start <= dt <= wk1_end:
            week = weeks[0]
        elif wk2_start <= dt <= wk2_end:
            week = weeks[1]
        else:
            continue

        if cat in weeks[0]:
            week[cat] += hours
            week["dates"].add(dt)
            week["has_data"] = True
            week["row_total"] += hours

    for w in weeks:
        sorted_dates = sorted(w["dates"])
        w["dates_str"] = ", ".join(format_date_short(d) for d in sorted_dates)

    return weeks


def _fmt_hours(h: float) -> str:
    if h == int(h):
        return f"{int(h)}.0"
    return f"{h:.1f}"


def merge_pdfs(pdf_bytes_list: List[bytes]) -> bytes:
    writer = PdfWriter()
    for pdf_bytes in pdf_bytes_list:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        writer.add_page(reader.pages[0])
    buf = io.BytesIO()
    writer.write(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def generate_ot_excel(employees_with_ot: list, pp_end: str) -> bytes:
    """Generate an Excel summary with stacked Wk1/Wk2 rows per employee."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Overtime Summary"

    end_dt = parse_date_flexible(pp_end)
    (wk1_start, wk1_end), (wk2_start, wk2_end) = pay_period_weeks(pp_end)

    title_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    bold = Font(bold=True, size=11)
    bold_sm = Font(bold=True, size=10)
    center = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    emp_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    total_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    ws.merge_cells("A1:G1")
    ws["A1"] = "City of Montebello — Transit Dept. 910 — Overtime Summary"
    ws["A1"].font = title_font

    ws["A2"] = f"Pay Period Ending: {end_dt.strftime('%m/%d/%Y')}"
    ws["A2"].font = bold

    ws["A3"] = f"Week 1: {wk1_start.strftime('%m/%d')} – {wk1_end.strftime('%m/%d/%Y')}"
    ws["A4"] = f"Week 2: {wk2_start.strftime('%m/%d')} – {wk2_end.strftime('%m/%d/%Y')}"

    headers = ["Employee", "Week", "OT 1.0", "OT 1.5", "CTE 1.0", "CTE 1.5", "Total"]
    header_row = 6
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_num = header_row + 1
    grand_total_all = 0.0
    cat_keys = ["ot10", "ot15", "cte10", "cte15"]

    for emp_ot in employees_with_ot:
        emp = emp_ot["employee"]
        ot_data = emp_ot["ot_data"]
        weeks = _aggregate_ot_by_week(ot_data, wk1_start, wk1_end, wk2_start, wk2_end)

        emp_total = sum(w["row_total"] for w in weeks)
        grand_total_all += emp_total
        emp_name = f"{emp['last']}, {emp['first']} (#{emp['emp_no']})"

        # Week 1 row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num + 1, end_column=1)
        name_cell = ws.cell(row=row_num, column=1, value=emp_name)
        name_cell.font = bold_sm
        name_cell.fill = emp_fill
        name_cell.border = thin_border
        name_cell.alignment = Alignment(vertical="center", wrap_text=True)

        wk1_total = weeks[0]["row_total"]
        wk1_dates_str = weeks[0]["dates_str"]
        wk1_label = f"Wk 1: {wk1_dates_str}" if wk1_dates_str else "Wk 1"
        ws.cell(row=row_num, column=2, value=wk1_label).font = bold_sm
        ws.cell(row=row_num, column=2).border = thin_border
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="left", wrap_text=True)
        for ci, ck in enumerate(cat_keys, 3):
            val = weeks[0][ck]
            cell = ws.cell(row=row_num, column=ci, value=val if val else "")
            cell.border = thin_border
            cell.alignment = center
        ws.cell(row=row_num, column=7, value=wk1_total if wk1_total else "")
        ws.cell(row=row_num, column=7).border = thin_border
        ws.cell(row=row_num, column=7).alignment = center

        row_num += 1

        # Week 2 row
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=1).fill = emp_fill
        wk2_total = weeks[1]["row_total"]
        wk2_dates_str = weeks[1]["dates_str"]
        wk2_label = f"Wk 2: {wk2_dates_str}" if wk2_dates_str else "Wk 2"
        ws.cell(row=row_num, column=2, value=wk2_label).font = bold_sm
        ws.cell(row=row_num, column=2).border = thin_border
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="left", wrap_text=True)
        for ci, ck in enumerate(cat_keys, 3):
            val = weeks[1][ck]
            cell = ws.cell(row=row_num, column=ci, value=val if val else "")
            cell.border = thin_border
            cell.alignment = center
        ws.cell(row=row_num, column=7, value=wk2_total if wk2_total else "")
        ws.cell(row=row_num, column=7).border = thin_border
        ws.cell(row=row_num, column=7).alignment = center

        row_num += 1

        # Employee total row
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
        total_label = ws.cell(row=row_num, column=1, value=f"Employee Total")
        total_label.font = bold_sm
        total_label.alignment = right_align
        total_label.border = thin_border
        for c in range(2, 7):
            ws.cell(row=row_num, column=c).border = thin_border
        total_val = ws.cell(row=row_num, column=7, value=emp_total)
        total_val.font = bold
        total_val.border = thin_border
        total_val.alignment = center

        row_num += 1

    # Grand total
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=6)
    gt_label = ws.cell(row=row_num, column=1, value="GRAND TOTAL")
    gt_label.font = Font(bold=True, size=12)
    gt_label.alignment = right_align
    gt_label.fill = total_fill
    gt_label.border = thin_border
    for c in range(2, 7):
        ws.cell(row=row_num, column=c).border = thin_border
        ws.cell(row=row_num, column=c).fill = total_fill
    gt_val = ws.cell(row=row_num, column=7, value=grand_total_all)
    gt_val.font = Font(bold=True, size=12)
    gt_val.border = thin_border
    gt_val.alignment = center
    gt_val.fill = total_fill

    col_widths = [30, 22, 10, 10, 10, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/parse-csv", methods=["POST"])
def parse_csv():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    f = request.files["file"]
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400

    file_bytes = f.read()
    try:
        employees = parse_employees_from_csv(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Failed to parse CSV: {e}"}), 400

    return jsonify({"employees": employees, "count": len(employees)})


@app.route("/api/generate-slips", methods=["POST"])
def generate_slips():
    """Feature 1: Generate pre-filled blank slips for all employees."""
    data = request.get_json()
    employees = data.get("employees", [])
    pp_end = data.get("payPeriodEnd", "")

    if not employees or not pp_end:
        return jsonify({"error": "Missing employees or pay period end date"}), 400

    employees.sort(key=lambda e: (e["last"].lower(), e["first"].lower()))

    pdf_list = []
    for emp in employees:
        try:
            pdf_bytes = fill_single_pdf(emp, pp_end)
            pdf_list.append(pdf_bytes)
        except Exception as e:
            app.logger.error(f"Error filling PDF for {emp}: {e}")
            continue

    if not pdf_list:
        return jsonify({"error": "No PDFs generated"}), 500

    merged = merge_pdfs(pdf_list)
    end_dt = parse_date_flexible(pp_end)

    return send_file(
        io.BytesIO(merged),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"Time_Exception_Slips_{end_dt.strftime('%m-%d-%y')}.pdf",
    )


@app.route("/api/generate-overtime", methods=["POST"])
def generate_overtime():
    """Feature 2: Generate OT-filled slips + Excel for employees with OT data."""
    data = request.get_json()
    employees_all = data.get("employees", [])
    pp_end = data.get("payPeriodEnd", "")
    ot_entries = data.get("otEntries", {})  # keyed by emp_no

    if not pp_end:
        return jsonify({"error": "Missing pay period end date"}), 400

    emps_with_ot = []
    for emp in employees_all:
        emp_no = emp.get("emp_no", "")
        if emp_no in ot_entries and ot_entries[emp_no].get("entries"):
            emps_with_ot.append({"employee": emp, "ot_data": ot_entries[emp_no]})

    if not emps_with_ot:
        return jsonify({"error": "No overtime entries found"}), 400

    emps_with_ot.sort(key=lambda x: (x["employee"]["last"].lower(), x["employee"]["first"].lower()))

    pdf_list = []
    for item in emps_with_ot:
        try:
            pdf_bytes = fill_single_pdf(item["employee"], pp_end, item["ot_data"])
            pdf_list.append(pdf_bytes)
        except Exception as e:
            app.logger.error(f"Error filling OT PDF for {item['employee']}: {e}")
            continue

    merged_pdf = merge_pdfs(pdf_list)
    excel_bytes = generate_ot_excel(emps_with_ot, pp_end)

    end_dt = parse_date_flexible(pp_end)
    date_str = end_dt.strftime("%m-%d-%y")

    return jsonify({
        "pdf": base64.b64encode(merged_pdf).decode("ascii"),
        "pdfFilename": f"Overtime_Slips_{date_str}.pdf",
        "excel": base64.b64encode(excel_bytes).decode("ascii"),
        "excelFilename": f"Overtime_Summary_{date_str}.xlsx",
    })


if __name__ == "__main__":
    app.run(debug=True, port=5050)
